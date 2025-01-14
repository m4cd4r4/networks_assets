import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
import openpyxl
from datetime import datetime
import os

# File path to your Excel file
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "network_assets.xlsx")

# Ensure the file exists and has necessary sheets
def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        # Create Models sheet
        models_sheet = wb.active
        models_sheet.title = "Models"
        models_sheet.append(["Model", "Serial #"])

        # Create Inventory sheet
        inventory_sheet = wb.create_sheet("Inventory")
        inventory_sheet.append(["Model", "Serial #"])

        # Create Timestamps sheet
        timestamps_sheet = wb.create_sheet("Timestamps")
        timestamps_sheet.append(["Timestamp", "Model", "Serial", "Action"])

        wb.save(EXCEL_FILE)

initialize_excel()

def log_action(model, serial, action):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ts_sheet = wb["Timestamps"]
    ts_sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), model, serial, action])
    wb.save(EXCEL_FILE)

def scan_serial(add_item=True):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    models_sheet = wb["Models"]
    inventory_sheet = wb["Inventory"]

    # Ensure Inventory has headers
    if inventory_sheet.max_row == 1 and inventory_sheet["A1"].value is None:
        inventory_sheet.append(["Model", "Serial #"])

    model = None
    serial_1 = simpledialog.askstring("Scan Serial", "Scan the model serial:")
    if not serial_1:
        messagebox.showerror("Error", "No serial scanned.")
        return

    serial_1 = serial_1.strip()  # Remove leading/trailing spaces

    # Look for the serial in the Models sheet
    for row in models_sheet.iter_rows(min_row=2, max_row=models_sheet.max_row):
        if str(row[1].value).strip() == serial_1:  # Match with Serial # column, ensuring proper formatting
            model = str(row[0].value).strip()  # Get the Model from the same row
            break

    if not model:
        messagebox.showerror("Error", "Model serial not found.")
        return

    # Add model to Inventory
    next_row = inventory_sheet.max_row + 1
    inventory_sheet.cell(row=next_row, column=1).value = model
    wb.save(EXCEL_FILE)

    # Log first action
    log_action(model, serial_1, "Add" if add_item else "Remove")

    serial_2 = simpledialog.askstring("Scan Serial", "Scan the unique serial:")
    if not serial_2:
        messagebox.showerror("Error", "No serial scanned.")
        return

    serial_2 = serial_2.strip()  # Remove leading/trailing spaces

    # Add unique serial to Inventory
    inventory_sheet.cell(row=next_row, column=2).value = serial_2
    wb.save(EXCEL_FILE)

    # Log second action
    log_action(model, serial_2, "Add" if add_item else "Remove")
    messagebox.showinfo("Success", f"Item ({model}) added to inventory.")

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Inventory Management")

        self.tree = ttk.Treeview(root, columns=("Model", "Serial"), show="headings")
        self.tree.heading("Model", text="Model")
        self.tree.heading("Serial", text="Serial")
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Buttons
        self.add_button = tk.Button(root, text="+", command=self.add_item, width=10)
        self.add_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.remove_button = tk.Button(root, text="-", command=self.remove_item, width=10)
        self.remove_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.refresh_button = tk.Button(root, text="Refresh", command=self.load_inventory, width=10)
        self.refresh_button.pack(side=tk.RIGHT, padx=5, pady=5)

        self.load_inventory()

    def load_inventory(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        inventory_sheet = wb["Inventory"]
        for row in inventory_sheet.iter_rows(min_row=2, max_row=inventory_sheet.max_row, values_only=True):
            self.tree.insert("", tk.END, values=row)

    def add_item(self):
        scan_serial(add_item=True)
        self.load_inventory()

    def remove_item(self):
        scan_serial(add_item=False)
        self.load_inventory()

if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()
