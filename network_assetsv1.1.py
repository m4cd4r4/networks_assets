import tkinter as tk
from tkinter import ttk, simpledialog, messagebox, Menu
import openpyxl
from datetime import datetime
import os
import subprocess
import sys
from contextlib import contextmanager

# File path to your Excel file
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "network_assets.xlsx")

def initialize_excel():
    """
    Initialize the Excel workbook.
    If the file does not exist, create it with the required sheets.
    If it exists, ensure that all required sheets are present.
    """
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        # Create Models sheet with headers
        models_sheet = wb.active
        models_sheet.title = "Models"
        models_sheet.append(["Model Serial", "Model Name"])  # First column: scanned model serial, second: linked model name

        # Create Inventory sheet with headers
        inventory_sheet = wb.create_sheet("Inventory")
        inventory_sheet.append(["Model", "Unique Serial"])

        # Create Timestamps sheet with headers
        timestamps_sheet = wb.create_sheet("Timestamps")
        timestamps_sheet.append(["Timestamp", "Model", "Serial", "Action"])

        wb.save(EXCEL_FILE)
    else:
        # If file exists, ensure each required sheet is present.
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Models" not in wb.sheetnames:
            ws = wb.create_sheet("Models")
            ws.append(["Model Serial", "Model Name"])
        if "Inventory" not in wb.sheetnames:
            ws = wb.create_sheet("Inventory")
            ws.append(["Model", "Unique Serial"])
        if "Timestamps" not in wb.sheetnames:
            ws = wb.create_sheet("Timestamps")
            ws.append(["Timestamp", "Model", "Serial", "Action"])
        wb.save(EXCEL_FILE)
        wb.close()

@contextmanager
def workbook_context(filepath):
    """
    Context manager to load and save the workbook.
    Ensures the workbook is saved and closed after operations.
    """
    wb = openpyxl.load_workbook(filepath)
    try:
        yield wb
    except Exception as e:
        messagebox.showerror("Error", f"Workbook error: {e}")
    finally:
        wb.save(filepath)
        wb.close()

def log_action(model, serial, action):
    """
    Log an action (Add or Remove) into the Timestamps sheet.
    """
    with workbook_context(EXCEL_FILE) as wb:
        ts_sheet = wb["Timestamps"]
        ts_sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), model, serial, action])

def open_spreadsheet():
    """
    Open the Excel spreadsheet using the default application.
    Cross-platform support is provided.
    """
    try:
        if sys.platform.startswith('win'):
            os.startfile(EXCEL_FILE)
        elif sys.platform.startswith('darwin'):
            subprocess.call(["open", EXCEL_FILE])
        else:
            subprocess.call(["xdg-open", EXCEL_FILE])
    except Exception as e:
        messagebox.showerror("Error", f"Unable to open spreadsheet: {e}")

def lookup_model(models_sheet, model_serial):
    """
    Look up the model name in the Models sheet using the scanned model serial.
    Returns the model name if found; otherwise, returns None.
    """
    model_serial = model_serial.strip().lower()
    for row in models_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and str(row[0]).strip().lower() == model_serial:
            return row[1]
    return None

def duplicate_exists(inventory_sheet, unique_serial):
    """
    Check if the unique serial already exists in the Inventory sheet.
    """
    unique_serial = unique_serial.strip().lower()
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        if row[1] is not None and str(row[1]).strip().lower() == unique_serial:
            return True
    return False

def add_inventory_item():
    """
    Adds a new item to the inventory by:
      1. Scanning a model serial and looking it up in the Models sheet.
      2. Scanning the unique serial.
      3. Checking for duplicates.
      4. Appending the item to the Inventory sheet and logging the action.
    """
    with workbook_context(EXCEL_FILE) as wb:
        models_sheet = wb["Models"]
        inventory_sheet = wb["Inventory"]

        # Scan model serial
        model_serial = simpledialog.askstring("Scan Serial", "Scan the model serial:")
        if not model_serial:
            messagebox.showinfo("Info", "Scanning cancelled.")
            return

        model_serial = model_serial.strip().lower()
        model_name = lookup_model(models_sheet, model_serial)
        if not model_name:
            messagebox.showerror("Error", f"Model serial not found: {model_serial}")
            return

        # Scan unique serial
        unique_serial = simpledialog.askstring("Scan Serial", "Scan the unique serial:")
        if not unique_serial:
            messagebox.showerror("Error", "No unique serial scanned.")
            return

        unique_serial = unique_serial.strip().lower()
        if duplicate_exists(inventory_sheet, unique_serial):
            messagebox.showerror("Error", "Duplicate unique serial found.")
            return

        # Append the new row to the Inventory sheet
        inventory_sheet.append([model_name, unique_serial])

    # Log the addition after workbook operations
    log_action(model_name, unique_serial, "Add")
    messagebox.showinfo("Success", f"Item ({model_name}) added to inventory.")

def remove_inventory_item():
    """
    Removes an item from the inventory by:
      1. Prompting for the unique serial.
      2. Searching the Inventory sheet for the matching row.
      3. Deleting the row if found and logging the action.
    """
    unique_serial = simpledialog.askstring("Remove Item", "Enter the unique serial to remove:")
    if not unique_serial:
        messagebox.showinfo("Info", "Removal cancelled.")
        return

    unique_serial = unique_serial.strip().lower()
    removed = False
    removed_model = None

    with workbook_context(EXCEL_FILE) as wb:
        inventory_sheet = wb["Inventory"]
        row_to_delete = None
        # Iterate over the rows starting from the second row (skip headers)
        for idx, row in enumerate(inventory_sheet.iter_rows(min_row=2), start=2):
            cell_value = row[1].value
            if cell_value is not None and str(cell_value).strip().lower() == unique_serial:
                row_to_delete = idx
                removed_model = row[0].value
                break

        if row_to_delete:
            inventory_sheet.delete_rows(row_to_delete)
            removed = True

    if removed:
        log_action(removed_model, unique_serial, "Remove")
        messagebox.showinfo("Success", f"Item with serial '{unique_serial}' removed.")
    else:
        messagebox.showerror("Error", f"Unique serial '{unique_serial}' not found in inventory.")

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Inventory Management")

        # Create the menu
        self.menu = Menu(self.root)
        self.root.config(menu=self.menu)

        # File menu
        file_menu = Menu(self.menu, tearoff=0)
        file_menu.add_command(label="Open Spreadsheet", command=open_spreadsheet)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        self.menu.add_cascade(label="File", menu=file_menu)

        # Create frames for layout using grid for more control
        self.frame_inventory = tk.Frame(root)
        self.frame_inventory.grid(row=0, column=0, sticky="nsew")
        self.frame_log = tk.Frame(root)
        self.frame_log.grid(row=0, column=1, sticky="nsew")

        # Configure grid weights for proper resizing
        root.grid_columnconfigure(0, weight=1)
        root.grid_columnconfigure(1, weight=1)
        root.grid_rowconfigure(0, weight=1)

        # Inventory TreeView
        self.tree = ttk.Treeview(self.frame_inventory, columns=("Model", "Serial"), show="headings")
        self.tree.heading("Model", text="Model")
        self.tree.heading("Serial", text="Serial")
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Log TreeView
        self.log_tree = ttk.Treeview(self.frame_log, columns=("Timestamp", "Model", "Serial", "Action"), show="headings")
        self.log_tree.heading("Timestamp", text="Timestamp")
        self.log_tree.heading("Model", text="Model")
        self.log_tree.heading("Serial", text="Serial")
        self.log_tree.heading("Action", text="Action")
        self.log_tree.pack(fill=tk.BOTH, expand=True)

        # Button frame at the bottom
        self.button_frame = tk.Frame(root)
        self.button_frame.grid(row=1, column=0, columnspan=2, pady=5)

        self.add_button = tk.Button(self.button_frame, text="+", command=self.handle_add, width=10)
        self.add_button.pack(side=tk.LEFT, padx=5)

        self.remove_button = tk.Button(self.button_frame, text="-", command=self.handle_remove, width=10)
        self.remove_button.pack(side=tk.LEFT, padx=5)

        self.refresh_button = tk.Button(self.button_frame, text="Refresh", command=self.load_data, width=10)
        self.refresh_button.pack(side=tk.RIGHT, padx=5)

        self.load_data()

    def load_data(self):
        """
        Load data from the Inventory and Timestamps sheets into the respective treeviews.
        """
        # Clear current data
        for row in self.tree.get_children():
            self.tree.delete(row)
        for row in self.log_tree.get_children():
            self.log_tree.delete(row)

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            inventory_sheet = wb["Inventory"]
            for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", tk.END, values=row)

            timestamps_sheet = wb["Timestamps"]
            for row in timestamps_sheet.iter_rows(min_row=2, values_only=True):
                self.log_tree.insert("", tk.END, values=row)
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Error loading data: {e}")

    def handle_add(self):
        """
        Callback for the '+' button.
        """
        add_inventory_item()
        self.load_data()

    def handle_remove(self):
        """
        Callback for the '-' button.
        """
        remove_inventory_item()
        self.load_data()

if __name__ == "__main__":
    initialize_excel()
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()
